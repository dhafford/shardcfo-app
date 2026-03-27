"""Core parsing engine — the main entry point for QBO export parsing.

Orchestrates the full pipeline:
  1. Detect report type and header metadata  (``detector``)
  2. Open the workbook and find column headers
  3. Parse every data row into a flat dict    (``parse_rows``)
  4. Reconstruct the section / row tree       (``build_tree``)
  5. Assemble the final ``ParsedReport``

The tree-building algorithm uses a **stack** to track parent–child
relationships derived from leading-space depth.  Total rows are attached
as the last child of their matching parent rather than as siblings.
"""

from __future__ import annotations

import re
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, FrozenSet, List, Optional, Tuple, Union

import openpyxl

from qbo_parser.detector import detect_report_type
from qbo_parser.models import (
    ParsedReport,
    ParseMetadata,
    ReportPeriod,
    Row,
    Section,
)
from qbo_parser.normalizer import (
    clean_amount,
    compute_depth,
    extract_account_code,
    is_total_row,
)
from qbo_parser.validator import validate_report


# ---------------------------------------------------------------------------
# Column-header detection
# ---------------------------------------------------------------------------

# "Jan 2024", "January 2024", "Dec 31, 2024", "December 31, 2024"
_MONTH_PARSE_FMTS = ("%b %Y", "%B %Y", "%b %d, %Y", "%B %d, %Y")
_QUARTER_RE = re.compile(r"^Q([1-4])\s+(\d{4})$", re.IGNORECASE)
_PCT_RE = re.compile(r"%\s*of\b", re.IGNORECASE)


def _format_column_header(value: Any) -> str:
    """Format a column-header cell value as a human-readable label.

    datetime(2019, 1, 1) → "Jan 2019"
    "Total"              → "Total"
    """
    if isinstance(value, datetime):
        return value.strftime("%b %Y")
    if isinstance(value, date):
        return value.strftime("%b %Y")
    if value is None:
        return ""
    return str(value).strip()


def is_percent_column(label: str) -> bool:
    """Return True if *label* looks like a QBO percentage column.

    Matches headers such as ``"% of Income"``, ``"% of Revenue"``,
    ``"% of Expenses"``, ``"% of Row"``.
    """
    stripped = label.strip()
    if stripped.startswith("%"):
        return True
    if _PCT_RE.search(stripped):
        return True
    return False


def _parse_column_date(label: str) -> Optional[date]:
    """Try to parse a column header as a month/quarter date.

    Returns the first-of-month date on success, None otherwise.
    """
    for fmt in _MONTH_PARSE_FMTS:
        try:
            return datetime.strptime(label, fmt).date()
        except ValueError:
            continue
    m = _QUARTER_RE.match(label)
    if m:
        quarter, year = int(m.group(1)), int(m.group(2))
        return date(year, (quarter - 1) * 3 + 1, 1)
    return None


def infer_period_from_columns(
    column_labels: List[str],
) -> Tuple[str, str]:
    """Derive a date range from column headers.

    Parses every label as a date (``"Jan 2024"`` → ``2024-01-01``),
    ignores non-date columns (``"Total"``, ``"% of Income"``), and
    returns the earliest start and latest end-of-month.

    Returns ``("", "")`` if no date columns are found.
    """
    dates = []  # type: List[date]
    for label in column_labels:
        d = _parse_column_date(label)
        if d is not None:
            dates.append(d)
    if not dates:
        return ("", "")
    start = min(dates)
    end_month = max(dates)
    _, last_day = monthrange(end_month.year, end_month.month)
    end = date(end_month.year, end_month.month, last_day)
    return (start.isoformat(), end.isoformat())


def _detect_columns(
    ws: Any,
    data_start_row: int,
    include_percent_cols: bool = False,
) -> Tuple[List[Tuple[int, str]], int]:
    """Identify column headers from the row at *data_start_row*.

    Parameters
    ----------
    ws : Worksheet
    data_start_row : int
        1-based row of the column header row.
    include_percent_cols : bool
        When False (default), columns whose header matches
        ``"% of Income"`` etc. are excluded.

    Returns
    -------
    data_columns : list of (0-based-col-index, label)
        One entry per retained column (skips column A).
    first_data_row : int
        1-based row number of the first account-data row.
    """
    for row in ws.iter_rows(min_row=data_start_row, max_row=data_start_row):
        cols = []  # type: List[Tuple[int, str]]
        for cell in row:
            if cell.column == 1:
                continue  # column A = labels
            if cell.value is not None:
                label = _format_column_header(cell.value)
                if not label:
                    continue
                if not include_percent_cols and is_percent_column(label):
                    continue
                cols.append((cell.column - 1, label))  # store 0-based
        return cols, data_start_row + 1
    return [], data_start_row + 1


# ---------------------------------------------------------------------------
# Flat row parsing
# ---------------------------------------------------------------------------

def _get_cell_depth(cell: Any) -> int:
    """Determine indent depth for a label cell.

    Priority:
      1. ``cell.alignment.indent`` (most reliable when QBO sets it)
      2. Leading whitespace in the cell value (the common case)
    """
    # Try alignment.indent first
    try:
        if cell.alignment and cell.alignment.indent:
            indent = int(cell.alignment.indent)
            if indent > 0:
                return indent
    except (AttributeError, TypeError):
        pass

    # Fall back to leading spaces (3 per level)
    if isinstance(cell.value, str):
        return compute_depth(cell.value)

    return 0


def parse_rows(
    ws: Any,
    first_data_row: int,
    data_columns: List[Tuple[int, str]],
) -> List[Dict]:
    """Parse every data row in the worksheet into a flat dict.

    Each returned dict has::

        {
            "account_name": str,
            "account_code": str,
            "depth":        int,
            "is_total":     bool,
            "amounts":      {col_label: float | None, …},
            "has_amounts":  bool,
        }

    Args:
        ws: The active openpyxl Worksheet.
        first_data_row: 1-based row where account data starts (after the
            column-header row).
        data_columns: ``(0-based-index, label)`` pairs from
            :func:`_detect_columns`.
    """
    flat = []  # type: List[Dict]

    # Limit column range for efficiency
    max_col_0 = max((idx for idx, _ in data_columns), default=0)

    for row_cells in ws.iter_rows(
        min_row=first_data_row,
        max_col=max_col_0 + 1,  # 1-based
    ):
        row = list(row_cells)
        if not row:
            continue

        label_cell = row[0]

        # Skip blank rows
        if label_cell.value is None:
            continue
        if isinstance(label_cell.value, str) and not label_cell.value.strip():
            continue

        raw_label = str(label_cell.value)
        depth = _get_cell_depth(label_cell)

        stripped = raw_label.strip()
        total = is_total_row(stripped)
        code, name = extract_account_code(stripped)

        # Collect amounts
        amounts = {}  # type: Dict[str, Optional[float]]
        has_any = False
        for col_idx, col_label in data_columns:
            val = clean_amount(row[col_idx].value) if col_idx < len(row) else None
            amounts[col_label] = val
            if val is not None:
                has_any = True

        flat.append({
            "account_name": name,
            "account_code": code,
            "depth": depth,
            "is_total": total,
            "amounts": amounts,
            "has_amounts": has_any,
        })

    return flat


# ---------------------------------------------------------------------------
# Tree building — rows within a section
# ---------------------------------------------------------------------------

def _build_row_tree(flat_rows: List[Dict]) -> List[Row]:
    """Reconstruct the nested Row tree from flat depth-annotated rows.

    Algorithm
    ---------
    Maintain a stack of ``(depth, Row, children_list)`` entries.

    * **Regular row at depth D** — pop the stack back to the nearest entry
      whose depth is *less than* D, append the new Row as a child, then
      push it onto the stack (it may have children of its own).

    * **Total row at depth D** — pop back the same way, then walk the
      parent's children list *backwards* to find the most recent non-total
      Row at depth D and attach the total as that Row's last child.  This
      handles QBO's convention where ``"   Total 61000 Advertising"``
      (depth 1) closes the group opened by ``"   61000 Advertising"``
      (also depth 1).
    """
    root_children = []  # type: List[Row]
    # (depth, Row-or-None, mutable-children-list)
    stack = [(-1, None, root_children)]  # type: List[Tuple[int, Optional[Row], List[Row]]]

    for rd in flat_rows:
        depth = rd["depth"]
        row = Row(
            account_name=rd["account_name"],
            account_code=rd["account_code"],
            depth=depth,
            amounts=rd["amounts"],
            is_total=rd["is_total"],
        )

        if rd["is_total"]:
            # Pop back to find the container for rows at this depth
            while len(stack) > 1 and stack[-1][0] >= depth:
                stack.pop()
            parent_children = stack[-1][2]

            # Attach to the last non-total Row at the same depth
            attached = False
            for i in range(len(parent_children) - 1, -1, -1):
                candidate = parent_children[i]
                if candidate.depth == depth and not candidate.is_total:
                    candidate.children.append(row)
                    attached = True
                    break
            if not attached:
                # Orphaned total — just add at this level
                parent_children.append(row)
        else:
            # Regular row — find parent (nearest stack entry with depth < D)
            while len(stack) > 1 and stack[-1][0] >= depth:
                stack.pop()
            parent_children = stack[-1][2]
            parent_children.append(row)
            # Push as potential parent for deeper rows
            stack.append((depth, row, row.children))

    return root_children


# ---------------------------------------------------------------------------
# Section grouping
# ---------------------------------------------------------------------------

_SECTION_HEADER_NAMES: FrozenSet[str] = frozenset(
    s.lower()
    for s in [
        # P&L
        "Income", "Revenue", "Cost of Goods Sold", "COGS",
        "Expenses", "Expense", "Operating Expenses",
        "Other Income", "Other Expenses", "Other Expense",
        # Balance Sheet
        "Assets", "Current Assets", "Fixed Assets", "Other Assets",
        "Non-Current Assets",
        "Liabilities", "Current Liabilities",
        "Long-Term Liabilities", "Long Term Liabilities",
        "Liabilities and Equity", "Liabilities & Equity",
        "Equity", "Stockholders Equity", "Stockholders' Equity",
        "Shareholders' Equity", "Owner's Equity",
        "Members' Equity", "Net Assets",
        # Cash Flow
        "Operating Activities", "Investing Activities",
        "Financing Activities",
        "Cash Flows from Operating Activities",
        "Cash Flows from Investing Activities",
        "Cash Flows from Financing Activities",
    ]
)


def _is_section_header(rd: Dict) -> bool:
    """Return True if a depth-0, non-total row is a section header.

    A section header is either:
      - a row with *no* amounts at all, or
      - a row whose name matches a known section label (safety net for
        rare cases where QBO puts a zero in a column).
    """
    if rd["depth"] != 0 or rd["is_total"]:
        return False
    if not rd["has_amounts"]:
        return True
    return rd["account_name"].lower().strip() in _SECTION_HEADER_NAMES


def build_tree(flat_rows: List[Dict]) -> List[Section]:
    """Organise flat rows into a list of :class:`Section` models.

    Splitting rules:

    * **Section header** — depth-0 row with no amounts (e.g. "Income",
      "Expenses").  Opens a new section.
    * **Section total** — depth-0 "Total …" row.  Closes the current
      section and records its amounts as ``Section.total``.
    * **Calculated row** — depth-0 row *with* amounts that is *not* a
      total (e.g. "Gross Profit", "Net Income").  Emitted as a
      standalone section with no child rows.
    * Everything else is collected into the current section's row list
      and later assembled into a nested tree by :func:`_build_row_tree`.
    """
    sections = []  # type: List[Section]
    cur_header = None  # type: Optional[Dict]
    cur_rows = []  # type: List[Dict]

    def _flush(total: Optional[Dict[str, Optional[float]]] = None) -> None:
        nonlocal cur_header, cur_rows
        if cur_header is not None:
            tree = _build_row_tree(cur_rows)
            sections.append(Section(
                name=cur_header["account_name"],
                rows=tree,
                total=total or {},
            ))
        cur_header = None
        cur_rows = []

    for rd in flat_rows:
        if rd["depth"] == 0:
            if rd["is_total"]:
                if cur_header is not None:
                    # Close the current section with this total
                    _flush(total=rd["amounts"])
                else:
                    # Orphaned total — no open section to close.
                    # Treat as standalone (e.g. "Gross Profit", "Net Income"
                    # are flagged as totals by the extended is_total_row but
                    # appear between sections without a matching header).
                    sections.append(Section(
                        name=rd["account_name"],
                        total=rd["amounts"],
                    ))
            elif _is_section_header(rd):
                _flush()
                cur_header = rd
            else:
                # Calculated / summary row with amounts, not a total, not
                # a section header.  Emit as standalone section.
                _flush()
                sections.append(Section(
                    name=rd["account_name"],
                    total=rd["amounts"],
                ))
        else:
            cur_rows.append(rd)

    _flush()  # close any trailing section
    return sections


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_qbo_report(
    filepath: Union[str, Path],
    include_percent_cols: bool = False,
) -> dict:
    """Parse a QuickBooks Online Excel export into structured data.

    This is the internal engine.  It reads an ``.xlsx`` file, detects the
    report type, extracts column headers (optionally including percentage
    columns), and builds a nested section / row hierarchy.

    Args:
        filepath: Path to the ``.xlsx`` file exported from QBO.
        include_percent_cols: When ``True``, keep ``"% of Income"``-style
            columns in the output.  When ``False`` (default), they are
            stripped so that only monetary amount columns remain.

    Returns:
        A JSON-serializable dict matching the :class:`ParsedReport` schema.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a valid ``.xlsx`` or has no data.
    """
    path = Path(filepath)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError(f"Expected an .xlsx file, got: {path.suffix}")
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    # 1. Detect report type and header metadata
    detection = detect_report_type(filepath)

    # 2. Open workbook (normal mode so cell.alignment is available)
    wb = openpyxl.load_workbook(str(path), data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ParsedReport(
                warnings=["Workbook has no active sheet"],
            ).model_dump(mode="json")

        # 3. Detect column headers
        data_columns, first_data_row = _detect_columns(
            ws,
            detection.data_start_row,
            include_percent_cols=include_percent_cols,
        )
        column_labels = [label for _, label in data_columns]

        if not data_columns:
            return ParsedReport(
                report_type=detection.report_type,
                company_name=detection.company_name,
                warnings=["No column headers detected"],
            ).model_dump(mode="json")

        # 4. Infer period from column headers (more accurate than header text)
        col_start, col_end = infer_period_from_columns(column_labels)
        period_start = col_start or detection.period_start
        period_end = col_end or detection.period_end

        # 5. Parse flat rows
        flat_rows = parse_rows(ws, first_data_row, data_columns)

        # 6. Build section/row tree
        section_list = build_tree(flat_rows)

        # 7. Assemble final report
        report_dict = ParsedReport(
            report_type=detection.report_type,
            company_name=detection.company_name,
            report_basis=detection.basis,
            period=ReportPeriod(
                start_date=period_start,
                end_date=period_end,
            ),
            columns=column_labels,
            sections=section_list,
            metadata=ParseMetadata(),
        ).model_dump(mode="json")

        # 8. Post-parse validation (adds warnings, never raises)
        report_dict["validation_warnings"] = validate_report(report_dict)

        return report_dict

    finally:
        wb.close()
