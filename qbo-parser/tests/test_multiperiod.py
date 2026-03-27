"""Tests for multi-period QBO exports with optional percentage columns.

Creates mock Excel files with openpyxl to exercise:
  - Multiple monthly amount columns + a "Total" column
  - "% of Income" percentage columns (stripped by default, included with flag)
  - Period inference from column-header dates (overrides header-text period)
  - Quarterly column headers
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from qbo_parser import parse_qbo_report, flatten_report
from qbo_parser.normalizer import detect_column_layout, is_percent_column
from qbo_parser.parser import (
    _detect_columns,
    _format_column_header,
    _parse_column_date,
    infer_period_from_columns,
    is_percent_column as parser_is_pct,
)


# ===================================================================
# Fixture — create a mock multi-period P&L with % columns
# ===================================================================


@pytest.fixture()
def multiperiod_xlsx(tmp_path) -> Path:
    """Create a small 3-month P&L with a Total and % of Income column."""
    fp = tmp_path / "multi_pnl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header block (rows 1-4)
    ws.cell(row=1, column=1, value="Acme Corp")
    ws.cell(row=2, column=1, value="Profit and Loss")
    ws.cell(row=3, column=1, value="January - March, 2024")
    # row 4 blank

    # Column headers (row 5): A empty, B-D months, E total, F percent
    ws.cell(row=5, column=2, value=datetime(2024, 1, 1))
    ws.cell(row=5, column=3, value=datetime(2024, 2, 1))
    ws.cell(row=5, column=4, value=datetime(2024, 3, 1))
    ws.cell(row=5, column=5, value="Total")
    ws.cell(row=5, column=6, value="% of Income")

    # --- Income ---
    ws.cell(row=6, column=1, value="Income")

    ws.cell(row=7, column=1, value="   41000 Sales")
    ws.cell(row=7, column=2, value=5000.0)
    ws.cell(row=7, column=3, value=6000.0)
    ws.cell(row=7, column=4, value=7000.0)
    ws.cell(row=7, column=5, value=18000.0)
    ws.cell(row=7, column=6, value="100.00%")

    ws.cell(row=8, column=1, value="   42000 Services")
    ws.cell(row=8, column=2, value=1000.0)
    ws.cell(row=8, column=3, value=1500.0)
    ws.cell(row=8, column=4, value=2000.0)
    ws.cell(row=8, column=5, value=4500.0)
    ws.cell(row=8, column=6, value="25.00%")

    ws.cell(row=9, column=1, value="Total Income")
    ws.cell(row=9, column=2, value=6000.0)
    ws.cell(row=9, column=3, value=7500.0)
    ws.cell(row=9, column=4, value=9000.0)
    ws.cell(row=9, column=5, value=22500.0)
    ws.cell(row=9, column=6, value="100.00%")

    # --- Expenses ---
    ws.cell(row=10, column=1, value="Expenses")

    ws.cell(row=11, column=1, value="   62000 Rent")
    ws.cell(row=11, column=2, value=1000.0)
    ws.cell(row=11, column=3, value=1000.0)
    ws.cell(row=11, column=4, value=1000.0)
    ws.cell(row=11, column=5, value=3000.0)
    ws.cell(row=11, column=6, value="13.33%")

    ws.cell(row=12, column=1, value="   63000 Utilities")
    ws.cell(row=12, column=2, value=200.0)
    ws.cell(row=12, column=3, value=250.0)
    ws.cell(row=12, column=4, value=300.0)
    ws.cell(row=12, column=5, value=750.0)
    ws.cell(row=12, column=6, value="3.33%")

    ws.cell(row=13, column=1, value="Total Expenses")
    ws.cell(row=13, column=2, value=1200.0)
    ws.cell(row=13, column=3, value=1250.0)
    ws.cell(row=13, column=4, value=1300.0)
    ws.cell(row=13, column=5, value=3750.0)
    ws.cell(row=13, column=6, value="16.67%")

    # --- Net Income ---
    ws.cell(row=14, column=1, value="Net Income")
    ws.cell(row=14, column=2, value=4800.0)
    ws.cell(row=14, column=3, value=6250.0)
    ws.cell(row=14, column=4, value=7700.0)
    ws.cell(row=14, column=5, value=18750.0)
    ws.cell(row=14, column=6, value="83.33%")

    wb.save(str(fp))
    wb.close()
    return fp


@pytest.fixture()
def quarterly_xlsx(tmp_path) -> Path:
    """Create a small quarterly P&L to test Q1/Q2 header parsing."""
    fp = tmp_path / "quarterly_pnl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Quarterly Corp")
    ws.cell(row=2, column=1, value="Profit and Loss")
    ws.cell(row=3, column=1, value="2024")
    # row 4 blank

    ws.cell(row=5, column=2, value="Q1 2024")
    ws.cell(row=5, column=3, value="Q2 2024")
    ws.cell(row=5, column=4, value="Total")

    ws.cell(row=6, column=1, value="Income")
    ws.cell(row=7, column=1, value="   Sales")
    ws.cell(row=7, column=2, value=30000.0)
    ws.cell(row=7, column=3, value=35000.0)
    ws.cell(row=7, column=4, value=65000.0)
    ws.cell(row=8, column=1, value="Total Income")
    ws.cell(row=8, column=2, value=30000.0)
    ws.cell(row=8, column=3, value=35000.0)
    ws.cell(row=8, column=4, value=65000.0)
    ws.cell(row=9, column=1, value="Net Income")
    ws.cell(row=9, column=2, value=30000.0)
    ws.cell(row=9, column=3, value=35000.0)
    ws.cell(row=9, column=4, value=65000.0)

    wb.save(str(fp))
    wb.close()
    return fp


# ===================================================================
# is_percent_column
# ===================================================================


class TestIsPercentColumn:
    def test_pct_of_income(self):
        assert is_percent_column("% of Income") is True

    def test_pct_of_revenue(self):
        assert is_percent_column("% of Revenue") is True

    def test_pct_of_expenses(self):
        assert is_percent_column("% of Expenses") is True

    def test_pct_of_row(self):
        assert is_percent_column("% of Row") is True

    def test_pct_bare(self):
        assert is_percent_column("%") is True

    def test_pct_with_spaces(self):
        assert is_percent_column("  % of Income  ") is True

    def test_not_pct_total(self):
        assert is_percent_column("Total") is False

    def test_not_pct_date(self):
        assert is_percent_column("Jan 2024") is False

    def test_not_pct_empty(self):
        assert is_percent_column("") is False

    def test_parser_module_agrees(self):
        """Both normalizer and parser versions should agree."""
        for label in ["% of Income", "Total", "Jan 2024", "%"]:
            assert is_percent_column(label) == parser_is_pct(label)


# ===================================================================
# _parse_column_date / infer_period_from_columns
# ===================================================================


class TestParseColumnDate:
    def test_abbreviated_month(self):
        from datetime import date
        assert _parse_column_date("Jan 2024") == date(2024, 1, 1)

    def test_full_month(self):
        from datetime import date
        assert _parse_column_date("December 2023") == date(2023, 12, 1)

    def test_quarter(self):
        from datetime import date
        assert _parse_column_date("Q1 2024") == date(2024, 1, 1)
        assert _parse_column_date("Q2 2024") == date(2024, 4, 1)
        assert _parse_column_date("Q3 2024") == date(2024, 7, 1)
        assert _parse_column_date("Q4 2024") == date(2024, 10, 1)

    def test_total_returns_none(self):
        assert _parse_column_date("Total") is None

    def test_pct_returns_none(self):
        assert _parse_column_date("% of Income") is None

    def test_empty(self):
        assert _parse_column_date("") is None


class TestInferPeriodFromColumns:
    def test_monthly_range(self):
        cols = ["Jan 2024", "Feb 2024", "Mar 2024", "Total"]
        start, end = infer_period_from_columns(cols)
        assert start == "2024-01-01"
        assert end == "2024-03-31"

    def test_full_year(self):
        cols = [f"{m} 2024" for m in
                ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]]
        cols.append("Total")
        start, end = infer_period_from_columns(cols)
        assert start == "2024-01-01"
        assert end == "2024-12-31"

    def test_quarterly(self):
        cols = ["Q1 2024", "Q2 2024", "Total"]
        start, end = infer_period_from_columns(cols)
        assert start == "2024-01-01"
        # Q2 → date(2024, 4, 1) → last day of April = 30
        assert end == "2024-04-30"

    def test_no_dates(self):
        start, end = infer_period_from_columns(["Total", "Budget"])
        assert start == ""
        assert end == ""

    def test_single_month(self):
        start, end = infer_period_from_columns(["Mar 2024", "Total"])
        assert start == "2024-03-01"
        assert end == "2024-03-31"

    def test_ignores_pct_columns(self):
        cols = ["Jan 2024", "% of Income", "Feb 2024"]
        start, end = infer_period_from_columns(cols)
        assert start == "2024-01-01"
        assert end == "2024-02-29"  # 2024 is a leap year

    def test_fixture_columns_override_header_text(self):
        """Our fixture says 'January - May 2019' but columns go through July."""
        result = parse_qbo_report("tests/fixtures/sample_pnl.xlsx")
        # Column-inferred period should cover all 7 months
        assert result["period"]["start_date"] == "2019-01-01"
        assert result["period"]["end_date"] == "2019-07-31"


# ===================================================================
# detect_column_layout — percent column classification
# ===================================================================


class TestDetectColumnLayoutPercent:
    def test_classifies_pct_column(self, multiperiod_xlsx):
        wb = openpyxl.load_workbook(str(multiperiod_xlsx), data_only=True)
        try:
            ws = wb.active
            layout = detect_column_layout(ws, header_row=5)

            assert layout["label_col"] == 0
            # Amount columns: B(1)=Jan, C(2)=Feb, D(3)=Mar, E(4)=Total
            assert len(layout["amount_cols"]) == 4
            assert layout["amount_cols"][1] == "Jan 2024"
            assert layout["amount_cols"][4] == "Total"

            # Percent column: F(5)=% of Income
            assert len(layout["percent_cols"]) == 1
            assert layout["percent_cols"][5] == "% of Income"
        finally:
            wb.close()


# ===================================================================
# _detect_columns — filtering percent columns
# ===================================================================


class TestDetectColumnsFiltering:
    def test_strips_pct_by_default(self, multiperiod_xlsx):
        wb = openpyxl.load_workbook(str(multiperiod_xlsx), data_only=True)
        try:
            ws = wb.active
            cols, _ = _detect_columns(ws, 5, include_percent_cols=False)
            labels = [label for _, label in cols]
            assert "% of Income" not in labels
            assert len(labels) == 4  # Jan, Feb, Mar, Total
        finally:
            wb.close()

    def test_includes_pct_when_flagged(self, multiperiod_xlsx):
        wb = openpyxl.load_workbook(str(multiperiod_xlsx), data_only=True)
        try:
            ws = wb.active
            cols, _ = _detect_columns(ws, 5, include_percent_cols=True)
            labels = [label for _, label in cols]
            assert "% of Income" in labels
            assert len(labels) == 5  # Jan, Feb, Mar, Total, % of Income
        finally:
            wb.close()


# ===================================================================
# Full parse — multi-period fixture
# ===================================================================


class TestParseMultiPeriod:
    def test_columns_detected(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        assert result["columns"] == ["Jan 2024", "Feb 2024", "Mar 2024", "Total"]

    def test_pct_column_stripped_by_default(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        assert "% of Income" not in result["columns"]

    def test_pct_column_included_with_flag(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx), include_percent_cols=True)
        assert "% of Income" in result["columns"]
        assert len(result["columns"]) == 5

    def test_period_inferred_from_columns(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        # Column-inferred dates override the header text
        assert result["period"]["start_date"] == "2024-01-01"
        assert result["period"]["end_date"] == "2024-03-31"

    def test_metadata(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        assert result["report_type"] == "profit_and_loss"
        assert result["company_name"] == "Acme Corp"

    def test_section_names(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        names = [s["name"] for s in result["sections"]]
        assert names == ["Income", "Expenses", "Net Income"]

    def test_amounts_per_period(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        income = result["sections"][0]
        sales = income["rows"][0]
        assert sales["account_name"] == "Sales"
        assert sales["amounts"]["Jan 2024"] == 5000.0
        assert sales["amounts"]["Feb 2024"] == 6000.0
        assert sales["amounts"]["Mar 2024"] == 7000.0
        assert sales["amounts"]["Total"] == 18000.0

    def test_section_total(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        income = result["sections"][0]
        assert income["total"]["Jan 2024"] == 6000.0
        assert income["total"]["Total"] == 22500.0

    def test_net_income(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        ni = result["sections"][-1]
        assert ni["name"] == "Net Income"
        assert ni["total"]["Jan 2024"] == 4800.0
        assert ni["total"]["Total"] == 18750.0

    def test_pct_amounts_when_included(self, multiperiod_xlsx):
        """When percent cols are included, the % values should parse as None
        (they're strings like '100.00%' which clean_amount can't parse)."""
        result = parse_qbo_report(
            str(multiperiod_xlsx), include_percent_cols=True,
        )
        income = result["sections"][0]
        sales = income["rows"][0]
        # "100.00%" can't be parsed by clean_amount → None
        assert sales["amounts"]["% of Income"] is None

    def test_flat_output(self, multiperiod_xlsx):
        result = parse_qbo_report(str(multiperiod_xlsx))
        flat = flatten_report(result)
        assert len(flat) > 0
        # Every row should have all 4 amount columns as keys
        for row in flat:
            assert "Jan 2024" in row
            assert "Total" in row
            assert "% of Income" not in row

    def test_flat_with_pct(self, multiperiod_xlsx):
        result = parse_qbo_report(
            str(multiperiod_xlsx), include_percent_cols=True,
        )
        flat = flatten_report(result)
        assert "% of Income" in flat[0]


# ===================================================================
# Full parse — quarterly fixture
# ===================================================================


class TestParseQuarterly:
    def test_quarterly_columns(self, quarterly_xlsx):
        result = parse_qbo_report(str(quarterly_xlsx))
        assert result["columns"] == ["Q1 2024", "Q2 2024", "Total"]

    def test_quarterly_period_inferred(self, quarterly_xlsx):
        result = parse_qbo_report(str(quarterly_xlsx))
        assert result["period"]["start_date"] == "2024-01-01"
        # Q2 starts April → last day of April = 30
        assert result["period"]["end_date"] == "2024-04-30"

    def test_quarterly_amounts(self, quarterly_xlsx):
        result = parse_qbo_report(str(quarterly_xlsx))
        income = result["sections"][0]
        sales = income["rows"][0]
        assert sales["amounts"]["Q1 2024"] == 30000.0
        assert sales["amounts"]["Q2 2024"] == 35000.0
        assert sales["amounts"]["Total"] == 65000.0
