"""Tests for post-parse accounting-identity validation.

Creates mock Balance Sheet, Cash Flow Statement, and intentionally-broken
fixtures to exercise every validator path.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from qbo_parser import parse_qbo_report
from qbo_parser.validator import (
    TOLERANCE,
    _find_section,
    _validate_balance_sheet,
    _validate_cash_flow,
    _validate_pnl,
    validate_report,
)


FIXTURES_DIR = Path(__file__).parent / "fixtures"


# ===================================================================
# Mock fixture builders
# ===================================================================


def _write_row(ws, row_num, label, *amounts):
    """Helper: write a label in col A and amounts in col B, C, …"""
    ws.cell(row=row_num, column=1, value=label)
    for i, val in enumerate(amounts, start=2):
        if val is not None:
            ws.cell(row=row_num, column=i, value=val)


@pytest.fixture()
def balance_sheet_xlsx(tmp_path) -> Path:
    """Mock Balance Sheet that balances: Assets 350k = Liab 30k + Equity 320k."""
    fp = tmp_path / "bs.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    _write_row(ws, 1, "Test Company")
    _write_row(ws, 2, "Balance Sheet")
    _write_row(ws, 3, "As of December 31, 2024")
    # row 4 blank
    ws.cell(row=5, column=2, value="Total")

    _write_row(ws, 6, "ASSETS")
    _write_row(ws, 7, "   Current Assets")
    _write_row(ws, 8, "      Cash and Cash Equivalents", 100000)
    _write_row(ws, 9, "      Accounts Receivable", 50000)
    _write_row(ws, 10, "   Total Current Assets", 150000)
    _write_row(ws, 11, "   Fixed Assets")
    _write_row(ws, 12, "      Property and Equipment", 200000)
    _write_row(ws, 13, "   Total Fixed Assets", 200000)
    _write_row(ws, 14, "TOTAL ASSETS", 350000)

    _write_row(ws, 15, "LIABILITIES AND EQUITY")
    _write_row(ws, 16, "   Liabilities")
    _write_row(ws, 17, "      Accounts Payable", 30000)
    _write_row(ws, 18, "   Total Liabilities", 30000)
    _write_row(ws, 19, "   Equity")
    _write_row(ws, 20, "      Retained Earnings", 320000)
    _write_row(ws, 21, "   Total Equity", 320000)
    _write_row(ws, 22, "TOTAL LIABILITIES AND EQUITY", 350000)

    wb.save(str(fp))
    wb.close()
    return fp


@pytest.fixture()
def balance_sheet_broken_xlsx(tmp_path) -> Path:
    """Mock BS where Assets 350k != Liab+Equity 345k → validation warning."""
    fp = tmp_path / "bs_broken.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    _write_row(ws, 1, "Broken Co")
    _write_row(ws, 2, "Balance Sheet")
    _write_row(ws, 3, "As of December 31, 2024")
    ws.cell(row=5, column=2, value="Total")

    _write_row(ws, 6, "ASSETS")
    _write_row(ws, 7, "   Cash", 350000)
    _write_row(ws, 8, "TOTAL ASSETS", 350000)

    _write_row(ws, 9, "LIABILITIES AND EQUITY")
    _write_row(ws, 10, "   Liabilities")
    _write_row(ws, 11, "      Accounts Payable", 30000)
    _write_row(ws, 12, "   Total Liabilities", 30000)
    _write_row(ws, 13, "   Equity")
    _write_row(ws, 14, "      Retained Earnings", 315000)
    _write_row(ws, 15, "   Total Equity", 315000)
    # Intentionally wrong: 30k + 315k = 345k ≠ 350k
    _write_row(ws, 16, "TOTAL LIABILITIES AND EQUITY", 345000)

    wb.save(str(fp))
    wb.close()
    return fp


@pytest.fixture()
def cash_flow_xlsx(tmp_path) -> Path:
    """Mock Statement of Cash Flows that reconciles."""
    fp = tmp_path / "scf.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    _write_row(ws, 1, "Test Company")
    _write_row(ws, 2, "Statement of Cash Flows")
    _write_row(ws, 3, "January - December, 2024")
    ws.cell(row=5, column=2, value="Total")

    _write_row(ws, 6, "OPERATING ACTIVITIES")
    _write_row(ws, 7, "   Net Income", 50000)
    _write_row(ws, 8, "   Depreciation and Amortization", 10000)
    _write_row(ws, 9, "   Changes in Accounts Receivable", -5000)
    _write_row(ws, 10, "Net cash provided by operating activities", 55000)

    _write_row(ws, 11, "INVESTING ACTIVITIES")
    _write_row(ws, 12, "   Purchase of Equipment", -20000)
    _write_row(ws, 13, "Net cash used in investing activities", -20000)

    _write_row(ws, 14, "FINANCING ACTIVITIES")
    _write_row(ws, 15, "   Loan Repayment", -10000)
    _write_row(ws, 16, "Net cash provided by financing activities", -10000)

    # 55000 + (-20000) + (-10000) = 25000
    _write_row(ws, 17, "NET CHANGE IN CASH", 25000)
    _write_row(ws, 18, "Cash at beginning of period", 100000)
    _write_row(ws, 19, "Cash at end of period", 125000)

    wb.save(str(fp))
    wb.close()
    return fp


@pytest.fixture()
def cash_flow_broken_xlsx(tmp_path) -> Path:
    """Mock SCF where activities don't sum to net change → warning."""
    fp = tmp_path / "scf_broken.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    _write_row(ws, 1, "Broken Co")
    _write_row(ws, 2, "Statement of Cash Flows")
    _write_row(ws, 3, "January - December, 2024")
    ws.cell(row=5, column=2, value="Total")

    _write_row(ws, 6, "OPERATING ACTIVITIES")
    _write_row(ws, 7, "   Net Income", 50000)
    _write_row(ws, 8, "Net cash provided by operating activities", 50000)

    _write_row(ws, 9, "INVESTING ACTIVITIES")
    _write_row(ws, 10, "   Purchases", -10000)
    _write_row(ws, 11, "Net cash used in investing activities", -10000)

    _write_row(ws, 12, "FINANCING ACTIVITIES")
    _write_row(ws, 13, "   Dividends", -5000)
    _write_row(ws, 14, "Net cash provided by financing activities", -5000)

    # 50000 - 10000 - 5000 = 35000, but we report 30000
    _write_row(ws, 15, "NET CHANGE IN CASH", 30000)

    wb.save(str(fp))
    wb.close()
    return fp


# ===================================================================
# _find_section helper
# ===================================================================


class TestFindSection:
    SECTIONS = [
        {"name": "Income", "total": {"T": 100}},
        {"name": "Other Income", "total": {"T": 10}},
        {"name": "Net Income", "total": {"T": 80}},
        {"name": "Net Operating Income", "total": {"T": 90}},
        {"name": "Cost of Goods Sold", "total": {"T": 20}},
    ]

    def test_exact_match(self):
        s = _find_section(self.SECTIONS, "income", exclude=["other", "net"])
        assert s["name"] == "Income"

    def test_exclude_filters(self):
        """'Other Income' and 'Net Income' should be excluded."""
        s = _find_section(self.SECTIONS, "income", exclude=["other", "net"])
        assert s["name"] == "Income"

    def test_find_other_income(self):
        s = _find_section(self.SECTIONS, "other income", exclude=["net"])
        assert s["name"] == "Other Income"

    def test_find_net_income(self):
        s = _find_section(self.SECTIONS, "net income", exclude=["other", "operating"])
        assert s["name"] == "Net Income"

    def test_find_cogs(self):
        s = _find_section(self.SECTIONS, "cost of goods")
        assert s["name"] == "Cost of Goods Sold"

    def test_not_found(self):
        assert _find_section(self.SECTIONS, "expenses") is None


# ===================================================================
# P&L validation (uses real fixture)
# ===================================================================


class TestValidatePnl:
    def test_fixture_passes(self):
        result = parse_qbo_report(str(FIXTURES_DIR / "sample_pnl.xlsx"))
        assert result["validation_warnings"] == []

    def test_validation_integrated(self):
        """validate_report is called automatically and stored in the dict."""
        result = parse_qbo_report(str(FIXTURES_DIR / "sample_pnl.xlsx"))
        assert "validation_warnings" in result
        assert isinstance(result["validation_warnings"], list)

    def test_unit_validator_passes(self):
        """Call the P&L validator directly with known-good data."""
        sections = [
            {"name": "Income", "total": {"T": 1000}},
            {"name": "Cost of Goods Sold", "total": {"T": 400}},
            {"name": "Expenses", "total": {"T": 300}},
            {"name": "Other Income", "total": {"T": 50}},
            {"name": "Other Expenses", "total": {"T": 100}},
            {"name": "Net Income", "total": {"T": 250}},  # 1000-400-300+50-100=250
        ]
        warnings = _validate_pnl(sections, ["T"])
        assert warnings == []

    def test_unit_validator_fails(self):
        sections = [
            {"name": "Income", "total": {"T": 1000}},
            {"name": "Cost of Goods Sold", "total": {"T": 400}},
            {"name": "Expenses", "total": {"T": 300}},
            {"name": "Net Income", "total": {"T": 999}},  # should be 300
        ]
        warnings = _validate_pnl(sections, ["T"])
        assert len(warnings) == 1
        assert "P&L" in warnings[0]
        assert "999" in warnings[0]

    def test_within_tolerance(self):
        """Diff <= TOLERANCE should NOT produce a warning."""
        sections = [
            {"name": "Income", "total": {"T": 1000}},
            {"name": "Expenses", "total": {"T": 700}},
            {"name": "Net Income", "total": {"T": 300.50}},  # off by 0.50
        ]
        warnings = _validate_pnl(sections, ["T"])
        assert warnings == []

    def test_missing_sections_no_crash(self):
        """If Income or Net Income is missing, return empty (can't validate)."""
        warnings = _validate_pnl([{"name": "Expenses", "total": {"T": 100}}], ["T"])
        assert warnings == []


# ===================================================================
# Balance Sheet parsing + validation
# ===================================================================


class TestBalanceSheet:
    def test_report_type(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assert result["report_type"] == "balance_sheet"

    def test_company_name(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assert result["company_name"] == "Test Company"

    def test_section_names(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        names = [s["name"] for s in result["sections"]]
        assert "ASSETS" in names
        assert "LIABILITIES AND EQUITY" in names

    def test_assets_total(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assets = next(s for s in result["sections"] if s["name"] == "ASSETS")
        assert assets["total"]["Total"] == 350000

    def test_liab_equity_total(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        le = next(s for s in result["sections"] if "LIABILITIES AND EQUITY" in s["name"])
        assert le["total"]["Total"] == 350000

    def test_nested_structure(self, balance_sheet_xlsx):
        """Current Assets and Fixed Assets should be sub-sections within ASSETS."""
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assets = next(s for s in result["sections"] if s["name"] == "ASSETS")
        row_names = [r["account_name"] for r in assets["rows"]]
        assert "Current Assets" in row_names
        assert "Fixed Assets" in row_names

    def test_sub_accounts_nested(self, balance_sheet_xlsx):
        """Cash should be nested under Current Assets."""
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assets = next(s for s in result["sections"] if s["name"] == "ASSETS")
        current = next(r for r in assets["rows"] if r["account_name"] == "Current Assets")
        child_names = [c["account_name"] for c in current["children"]]
        assert "Cash and Cash Equivalents" in child_names
        assert "Accounts Receivable" in child_names

    def test_validation_passes(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assert result["validation_warnings"] == []

    def test_validation_fails_when_unbalanced(self, balance_sheet_broken_xlsx):
        result = parse_qbo_report(str(balance_sheet_broken_xlsx))
        assert len(result["validation_warnings"]) == 1
        assert "BS" in result["validation_warnings"][0]
        assert "350,000" in result["validation_warnings"][0]
        assert "345,000" in result["validation_warnings"][0]

    def test_bs_period_as_of(self, balance_sheet_xlsx):
        result = parse_qbo_report(str(balance_sheet_xlsx))
        assert result["period"]["start_date"] == ""
        assert result["period"]["end_date"] == "2024-12-31"


# ===================================================================
# Cash Flow Statement parsing + validation
# ===================================================================


class TestCashFlowStatement:
    def test_report_type(self, cash_flow_xlsx):
        result = parse_qbo_report(str(cash_flow_xlsx))
        assert result["report_type"] == "cash_flow_statement"

    def test_section_names(self, cash_flow_xlsx):
        result = parse_qbo_report(str(cash_flow_xlsx))
        names = [s["name"] for s in result["sections"]]
        assert any("OPERATING" in n.upper() for n in names)
        assert any("INVESTING" in n.upper() for n in names)
        assert any("FINANCING" in n.upper() for n in names)

    def test_operating_total(self, cash_flow_xlsx):
        result = parse_qbo_report(str(cash_flow_xlsx))
        operating = next(
            s for s in result["sections"]
            if "operating" in s["name"].lower()
        )
        assert operating["total"]["Total"] == 55000

    def test_net_change_in_cash(self, cash_flow_xlsx):
        result = parse_qbo_report(str(cash_flow_xlsx))
        nc = next(
            s for s in result["sections"]
            if "net change" in s["name"].lower()
        )
        assert nc["total"]["Total"] == 25000

    def test_cash_beginning_end(self, cash_flow_xlsx):
        """Cash at beginning/end of period should appear as standalone sections."""
        result = parse_qbo_report(str(cash_flow_xlsx))
        names = [s["name"].lower() for s in result["sections"]]
        assert any("beginning" in n for n in names)
        assert any("end" in n for n in names)

    def test_cash_end_value(self, cash_flow_xlsx):
        result = parse_qbo_report(str(cash_flow_xlsx))
        end = next(
            s for s in result["sections"]
            if "end" in s["name"].lower()
        )
        assert end["total"]["Total"] == 125000

    def test_validation_passes(self, cash_flow_xlsx):
        result = parse_qbo_report(str(cash_flow_xlsx))
        assert result["validation_warnings"] == []

    def test_validation_fails_when_mismatch(self, cash_flow_broken_xlsx):
        result = parse_qbo_report(str(cash_flow_broken_xlsx))
        assert len(result["validation_warnings"]) == 1
        assert "SCF" in result["validation_warnings"][0]

    def test_operating_rows(self, cash_flow_xlsx):
        """Operating Activities should contain Net Income and adjustments."""
        result = parse_qbo_report(str(cash_flow_xlsx))
        operating = next(
            s for s in result["sections"]
            if "operating" in s["name"].lower()
        )
        row_names = [r["account_name"] for r in operating["rows"]]
        assert "Net Income" in row_names or any("net income" in n.lower() for n in row_names)


# ===================================================================
# validate_report dispatcher
# ===================================================================


class TestValidateReport:
    def test_unknown_type_returns_empty(self):
        assert validate_report({"report_type": "unknown", "sections": [], "columns": []}) == []

    def test_no_sections_returns_empty(self):
        assert validate_report({"report_type": "profit_and_loss", "sections": [], "columns": ["T"]}) == []

    def test_dispatches_to_pnl(self):
        report = {
            "report_type": "profit_and_loss",
            "columns": ["T"],
            "sections": [
                {"name": "Income", "total": {"T": 100}},
                {"name": "Net Income", "total": {"T": 100}},
            ],
        }
        # Should pass (100 - 0 - 0 + 0 - 0 = 100)
        assert validate_report(report) == []

    def test_dispatches_to_bs(self):
        report = {
            "report_type": "balance_sheet",
            "columns": ["T"],
            "sections": [
                {"name": "ASSETS", "total": {"T": 100}},
                {"name": "LIABILITIES AND EQUITY", "total": {"T": 100}},
            ],
        }
        assert validate_report(report) == []

    def test_dispatches_to_scf(self):
        report = {
            "report_type": "cash_flow_statement",
            "columns": ["T"],
            "sections": [
                {"name": "Operating Activities", "total": {"T": 50}},
                {"name": "Investing Activities", "total": {"T": -20}},
                {"name": "Financing Activities", "total": {"T": -5}},
                {"name": "Net Change in Cash", "total": {"T": 25}},
            ],
        }
        assert validate_report(report) == []
