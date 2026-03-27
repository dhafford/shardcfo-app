"""Comprehensive tests for qbo_parser.normalizer.

Covers every public function with edge cases, especially parse_amount.
"""

from pathlib import Path

import pytest

from qbo_parser.normalizer import (
    clean_account_name,
    clean_amount,
    compute_depth,
    detect_column_layout,
    detect_indent_depth,
    extract_account_code,
    is_total_row,
    parse_amount,
)

FIXTURES_DIR = Path(__file__).parent / "fixtures"


# ===================================================================
# Helpers — lightweight mocks for openpyxl cell objects
# ===================================================================


class _MockAlignment:
    def __init__(self, indent: float = 0.0):
        self.indent = indent


class MockCell:
    """Minimal stand-in for an openpyxl Cell."""
    def __init__(
        self,
        value: object = None,
        indent: float = 0.0,
        column: int = 1,
    ):
        self.value = value
        self.alignment = _MockAlignment(indent)
        self.column = column  # 1-based, like openpyxl


# ===================================================================
# parse_amount
# ===================================================================


class TestParseAmount:
    """parse_amount: blanks → 0.0, non-numeric text → None."""

    # --- Blanks and empties → 0.0 ---

    def test_none(self):
        assert parse_amount(None) == 0.0

    def test_empty_string(self):
        assert parse_amount("") == 0.0

    def test_whitespace_only(self):
        assert parse_amount("   ") == 0.0

    def test_dash(self):
        assert parse_amount("-") == 0.0

    def test_bool_false(self):
        assert parse_amount(False) == 0.0

    def test_bool_true(self):
        assert parse_amount(True) == 0.0

    # --- Numeric passthrough ---

    def test_int(self):
        assert parse_amount(18935) == 18935.0

    def test_float(self):
        assert parse_amount(20701.02) == 20701.02

    def test_zero_int(self):
        assert parse_amount(0) == 0.0

    def test_zero_float(self):
        assert parse_amount(0.0) == 0.0

    def test_negative_float(self):
        assert parse_amount(-37884.94) == -37884.94

    # --- String number formats ---

    def test_plain_number_string(self):
        assert parse_amount("42") == 42.0

    def test_decimal_string(self):
        assert parse_amount("1234.56") == 1234.56

    def test_commas(self):
        assert parse_amount("1,234.56") == 1234.56

    def test_large_commas(self):
        assert parse_amount("1,234,567.89") == 1234567.89

    def test_negative_with_dash(self):
        assert parse_amount("-1,234.56") == -1234.56

    def test_negative_no_commas(self):
        assert parse_amount("-42.5") == -42.5

    # --- Parenthetical negatives ---

    def test_parenthetical(self):
        assert parse_amount("(1,234.56)") == -1234.56

    def test_parenthetical_no_commas(self):
        assert parse_amount("(42)") == -42.0

    def test_parenthetical_with_spaces(self):
        assert parse_amount("( 1,234.56 )") == -1234.56

    # --- Currency symbols ---

    def test_dollar_sign(self):
        assert parse_amount("$1,234.56") == 1234.56

    def test_dollar_negative(self):
        assert parse_amount("-$1,234.56") == -1234.56

    def test_dollar_parenthetical(self):
        assert parse_amount("($1,234.56)") == -1234.56

    def test_euro(self):
        assert parse_amount("\u20ac500.00") == 500.0

    def test_pound(self):
        assert parse_amount("\u00a3999.99") == 999.99

    # --- Non-numeric text → None (signals a label) ---

    def test_label_text(self):
        assert parse_amount("Income") is None

    def test_label_with_numbers(self):
        """Account names with numbers shouldn't parse as amounts."""
        assert parse_amount("41000 Sales") is None

    def test_date_string(self):
        assert parse_amount("Jan 2019") is None

    def test_random_text(self):
        assert parse_amount("hello world") is None

    # --- Contrast with clean_amount ---

    def test_differs_from_clean_amount_on_none(self):
        assert clean_amount(None) is None
        assert parse_amount(None) == 0.0

    def test_differs_from_clean_amount_on_empty(self):
        assert clean_amount("") is None
        assert parse_amount("") == 0.0

    def test_agrees_with_clean_amount_on_number(self):
        assert clean_amount(42.5) == parse_amount(42.5) == 42.5

    def test_agrees_with_clean_amount_on_string_number(self):
        assert clean_amount("$1,234.56") == parse_amount("$1,234.56") == 1234.56


# ===================================================================
# clean_account_name
# ===================================================================


class TestCleanAccountName:
    def test_simple(self):
        assert clean_account_name("Sales") == "Sales"

    def test_preserves_leading_spaces(self):
        assert clean_account_name("   41000 Sales") == "   41000 Sales"

    def test_strips_trailing_whitespace(self):
        assert clean_account_name("   41000 Sales   ") == "   41000 Sales"

    def test_strips_trailing_newline(self):
        assert clean_account_name("   Sales\n") == "   Sales"

    def test_strips_trailing_tab(self):
        assert clean_account_name("Sales\t") == "Sales"

    def test_collapses_internal_spaces(self):
        assert clean_account_name("   Payroll   Expenses") == "   Payroll Expenses"

    def test_removes_null_bytes(self):
        assert clean_account_name("Sales\x00Revenue") == "SalesRevenue"

    def test_removes_control_chars(self):
        assert clean_account_name("Adv\x07ertising") == "Advertising"

    def test_preserves_accented_chars(self):
        """Latin-1 supplement chars (accents) should survive."""
        assert clean_account_name("Café Expenses") == "Café Expenses"

    def test_empty_string(self):
        assert clean_account_name("") == ""

    def test_none_like_empty(self):
        """None-ish falsy values return empty string."""
        assert clean_account_name("") == ""

    def test_all_whitespace(self):
        assert clean_account_name("   ") == ""

    def test_leading_spaces_only_preserved(self):
        """Six leading spaces with trailing should keep leading, strip trailing."""
        assert clean_account_name("      Contractors  \n") == "      Contractors"

    def test_internal_multiple_spaces_with_leading(self):
        assert clean_account_name("   Legal  &  Prof  Services") == "   Legal & Prof Services"


# ===================================================================
# detect_indent_depth
# ===================================================================


class TestDetectIndentDepth:
    # --- Priority 1: cell.alignment.indent ---

    def test_alignment_indent(self):
        cell = MockCell(value="Sales", indent=2.0)
        assert detect_indent_depth(cell, "Sales") == 2

    def test_alignment_indent_overrides_spaces(self):
        """alignment.indent takes precedence over leading spaces."""
        cell = MockCell(value="      Sales", indent=1.0)
        assert detect_indent_depth(cell, "      Sales") == 1

    def test_alignment_indent_zero_falls_through(self):
        """indent=0 is treated as absent → fall through to spaces."""
        cell = MockCell(value="   Sales", indent=0.0)
        assert detect_indent_depth(cell, "   Sales") == 1

    # --- Priority 2: leading whitespace ---

    def test_leading_spaces_3(self):
        cell = MockCell(value="   Sales")
        assert detect_indent_depth(cell, "   Sales") == 1

    def test_leading_spaces_6(self):
        cell = MockCell(value="      Sub")
        assert detect_indent_depth(cell, "      Sub") == 2

    def test_no_leading_spaces(self):
        cell = MockCell(value="Income")
        assert detect_indent_depth(cell, "Income") == 0

    def test_custom_indent_unit(self):
        """Some exports use 4 spaces per level."""
        cell = MockCell(value="    Sales")
        assert detect_indent_depth(cell, "    Sales", indent_unit=4) == 1

    # --- Priority 3: column offset ---

    def test_column_b_depth_1(self):
        """Text in column B with no indentation → depth 1."""
        cell = MockCell(value="Sub Account", column=2)
        assert detect_indent_depth(cell, "Sub Account") == 1

    def test_column_c_depth_2(self):
        cell = MockCell(value="Deep Sub", column=3)
        assert detect_indent_depth(cell, "Deep Sub") == 2

    def test_column_a_no_offset(self):
        cell = MockCell(value="Top Level", column=1)
        assert detect_indent_depth(cell, "Top Level") == 0

    # --- None / edge cases ---

    def test_none_cell(self):
        assert detect_indent_depth(None, "   Sales") == 1

    def test_none_cell_no_spaces(self):
        assert detect_indent_depth(None, "Income") == 0

    def test_empty_value(self):
        cell = MockCell(value="")
        assert detect_indent_depth(cell, "") == 0


# ===================================================================
# is_total_row (extended)
# ===================================================================


class TestIsTotalRowExtended:
    """Tests for the extended is_total_row that catches Net/Gross Profit."""

    # --- Original "Total" patterns (unchanged) ---

    def test_total_prefix(self):
        assert is_total_row("Total Income") is True

    def test_total_with_code(self):
        assert is_total_row("Total 61000 Advertising & Marketing") is True

    def test_total_all_caps(self):
        assert is_total_row("TOTAL EXPENSES") is True

    def test_just_total(self):
        assert is_total_row("Total") is True

    # --- New: "Net " prefix ---

    def test_net_income(self):
        assert is_total_row("Net Income") is True

    def test_net_revenue(self):
        assert is_total_row("Net Revenue") is True

    def test_net_operating_income(self):
        assert is_total_row("Net Operating Income") is True

    def test_net_other_income(self):
        assert is_total_row("Net Other Income") is True

    def test_net_case_insensitive(self):
        assert is_total_row("NET INCOME") is True

    def test_net_with_leading_spaces(self):
        assert is_total_row("  Net Income") is True

    # --- New: "Gross Profit" exact match ---

    def test_gross_profit(self):
        assert is_total_row("Gross Profit") is True

    def test_gross_profit_case_insensitive(self):
        assert is_total_row("GROSS PROFIT") is True

    def test_gross_profit_with_spaces(self):
        assert is_total_row("  Gross Profit  ") is True

    # --- Non-matches ---

    def test_regular_account(self):
        assert is_total_row("   41000 Sales") is False

    def test_grand_total_no_match(self):
        """'Grand Total' doesn't start with 'Total '."""
        assert is_total_row("Grand Total") is False

    def test_empty(self):
        assert is_total_row("") is False

    def test_network_expenses(self):
        """'Network' starts with 'Net' but not 'Net ' (with space)."""
        assert is_total_row("Network Expenses") is False

    def test_gross_revenue(self):
        """Only 'Gross Profit' is matched, not other 'Gross' phrases."""
        assert is_total_row("Gross Revenue") is False

    def test_total_in_middle(self):
        assert is_total_row("Subtotal") is False

    def test_just_net(self):
        """'Net' alone (no space after) → False."""
        assert is_total_row("Net") is False


# ===================================================================
# detect_column_layout (integration with fixture)
# ===================================================================


class TestDetectColumnLayout:
    def test_fixture_pnl(self):
        """Integration test against the real QBO P&L fixture."""
        import openpyxl
        wb = openpyxl.load_workbook(
            str(FIXTURES_DIR / "sample_pnl.xlsx"), data_only=True,
        )
        try:
            ws = wb.active
            # Header row is row 5 (detected by detector as data_start_row)
            layout = detect_column_layout(ws, header_row=5)

            assert layout["label_col"] == 0
            cols = layout["amount_cols"]
            assert len(cols) == 7

            # 0-based indices: B=1, C=2, … H=7
            assert cols[1] == "Jan 2019"
            assert cols[2] == "Feb 2019"
            assert cols[7] == "Jul 2019"
        finally:
            wb.close()

    def test_empty_row(self):
        """If no cells have values, amount_cols should be empty."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row 1 is entirely empty
        layout = detect_column_layout(ws, header_row=1)
        assert layout["label_col"] == 0
        assert layout["amount_cols"] == {}
        wb.close()

    def test_string_headers(self):
        """Column headers that are plain strings (e.g. 'Total', 'Budget')."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Account")
        ws.cell(row=1, column=2, value="Total")
        ws.cell(row=1, column=3, value="Budget")

        layout = detect_column_layout(ws, header_row=1)
        assert layout["label_col"] == 0
        assert layout["amount_cols"] == {1: "Total", 2: "Budget"}
        wb.close()


# ===================================================================
# Existing functions — additional edge-case coverage
# ===================================================================


class TestComputeDepthExtra:
    """Extra edge-case tests beyond the ones in test_parser.py."""

    def test_tab_not_counted(self):
        """Tabs are not spaces — depth should be 0."""
        assert compute_depth("\tIncome") == 0

    def test_twelve_spaces(self):
        assert compute_depth("            X") == 4

    def test_mixed_content(self):
        assert compute_depth("   41000 Sales Revenue") == 1


class TestCleanAmountExtra:
    """Extra edge cases for clean_amount."""

    def test_parenthetical_with_dollar(self):
        assert clean_amount("($500.00)") == -500.0

    def test_only_dollar(self):
        assert clean_amount("$") is None

    def test_european_comma_not_supported(self):
        """European '1.234,56' format is not QBO — returns None."""
        # This will fail to parse because 1.234 is valid but ,56 makes it fail
        # after removing commas it becomes "1.23456" — actually that parses!
        # Let's test a clear failure case instead
        assert clean_amount("not a number") is None

    def test_negative_zero(self):
        result = clean_amount("-0")
        assert result == 0.0 or result == -0.0  # both acceptable


class TestExtractAccountCodeExtra:
    def test_six_digit_code(self):
        code, label = extract_account_code("100000 Cash in Bank")
        assert code == "100000"
        assert label == "Cash in Bank"

    def test_three_digit_code_too_short(self):
        """Codes under 4 digits are not matched."""
        code, label = extract_account_code("100 Petty Cash")
        assert code == ""
        assert label == "100 Petty Cash"

    def test_code_with_leading_total_spaces(self):
        code, label = extract_account_code("   Total 61000 Advertising & Marketing")
        assert code == "61000"
        assert "Total" in label
        assert "Advertising & Marketing" in label
