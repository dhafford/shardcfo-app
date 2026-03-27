"""Generate tests/fixtures/sample_scf.xlsx — Statement of Cash Flows for Demo Startup Inc."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

FIXTURES_DIR = Path(__file__).parent
OUTPUT_PATH = FIXTURES_DIR / "sample_scf.xlsx"

BOLD = Font(bold=True)


def _write_header(ws, company: str, title: str, date_range: str) -> None:
    """Write the 3-row QBO header block."""
    ws["A1"] = company
    ws["A1"].font = BOLD
    ws["A2"] = title
    ws["A2"].font = BOLD
    ws["A3"] = date_range
    ws["A3"].font = BOLD
    ws["B5"] = "Total"
    ws["B5"].font = BOLD


def _write_row(ws, row_num: int, label: str, amount=None, bold_amount: bool = False) -> None:
    """Write a single data row.  Label is always bold (QBO convention)."""
    cell_a = ws.cell(row=row_num, column=1, value=label)
    cell_a.font = BOLD
    if amount is not None:
        cell_b = ws.cell(row=row_num, column=2, value=float(amount))
        if bold_amount:
            cell_b.font = BOLD


def generate() -> Path:
    """Create sample_scf.xlsx and return its Path.

    Cash flow identities:
      Operating (34400) + Investing (-45000) + Financing (55000) = 44400
      Beginning (240600) + Net Change (44400) = End (285000)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_header(
        ws,
        company="Demo Startup Inc",
        title="Statement of Cash Flows",
        date_range="January - December, 2024",
    )

    row = 6  # first data row

    # --- OPERATING ACTIVITIES ---
    ws.cell(row=row, column=1, value="OPERATING ACTIVITIES").font = BOLD
    row += 1

    _write_row(ws, row, "   Net Income", 5400)
    row += 1
    _write_row(ws, row, "   Adjustments to reconcile Net Income")
    row += 1
    _write_row(ws, row, "      Depreciation and Amortization", 12000)
    row += 1
    _write_row(ws, row, "      Changes in Accounts Receivable", -15000)
    row += 1
    _write_row(ws, row, "      Changes in Prepaid Expenses", -3000)
    row += 1
    _write_row(ws, row, "      Changes in Accounts Payable", 8000)
    row += 1
    _write_row(ws, row, "      Changes in Accrued Expenses", 5000)
    row += 1
    _write_row(ws, row, "      Changes in Deferred Revenue", 22000)
    row += 1
    _write_row(ws, row, "   Total Adjustments", 29000, bold_amount=True)
    row += 1

    # Depth-0 summary: net cash from operating
    _write_row(ws, row, "Net cash provided by operating activities", 34400, bold_amount=True)
    row += 1

    # --- INVESTING ACTIVITIES ---
    ws.cell(row=row, column=1, value="INVESTING ACTIVITIES").font = BOLD
    row += 1

    _write_row(ws, row, "   Purchase of Equipment", -45000)
    row += 1

    # Depth-0 summary
    _write_row(ws, row, "Net cash used in investing activities", -45000, bold_amount=True)
    row += 1

    # --- FINANCING ACTIVITIES ---
    ws.cell(row=row, column=1, value="FINANCING ACTIVITIES").font = BOLD
    row += 1

    _write_row(ws, row, "   Proceeds from Notes Payable", 50000)
    row += 1
    _write_row(ws, row, "   Repayment of Notes Payable", -20000)
    row += 1
    _write_row(ws, row, "   Issuance of Common Stock", 25000)
    row += 1

    # Depth-0 summary
    _write_row(ws, row, "Net cash provided by financing activities", 55000, bold_amount=True)
    row += 1

    # --- NET CHANGE IN CASH ---
    _write_row(ws, row, "NET CHANGE IN CASH", 44400, bold_amount=True)
    row += 1

    # --- Beginning / ending cash (standalone depth-0 rows with amounts) ---
    _write_row(ws, row, "Cash at beginning of period", 240600, bold_amount=True)
    row += 1
    _write_row(ws, row, "Cash at end of period", 285000, bold_amount=True)

    wb.save(str(OUTPUT_PATH))
    print(f"Generated: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    generate()
