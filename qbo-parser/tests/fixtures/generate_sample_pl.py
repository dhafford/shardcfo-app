"""Generate tests/fixtures/sample_pl.xlsx — Annual P&L for Demo Startup Inc."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

FIXTURES_DIR = Path(__file__).parent
OUTPUT_PATH = FIXTURES_DIR / "sample_pl.xlsx"

BOLD = Font(bold=True)


def _write_header(ws, company: str, title: str, date_range: str) -> None:
    """Write the 3-row QBO header block."""
    ws["A1"] = company
    ws["A1"].font = BOLD
    ws["A2"] = title
    ws["A2"].font = BOLD
    ws["A3"] = date_range
    ws["A3"].font = BOLD
    # Row 4 blank, Row 5 is column header row
    ws["B5"] = "Total"
    ws["B5"].font = BOLD


def _write_row(ws, row_num: int, label: str, amount=None, bold_amount: bool = False) -> None:
    """Write a single data row.  Label is always bold (QBO convention)."""
    cell_a = ws.cell(row=row_num, column=1, value=label)
    cell_a.font = BOLD
    if amount is not None:
        cell_b = ws.cell(row=row_num, column=2, value=float(amount))
        if bold_amount:
            cell_b.font = BOLD


def generate() -> Path:
    """Create sample_pl.xlsx and return its Path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_header(
        ws,
        company="Demo Startup Inc",
        title="Profit and Loss (Accrual Basis)",
        date_range="January - December, 2024",
    )

    row = 6  # first data row

    # --- Income ---
    ws.cell(row=row, column=1, value="Income").font = BOLD
    row += 1

    _write_row(ws, row, "   40000 SaaS Revenue", 720000)
    row += 1
    _write_row(ws, row, "   40100 Professional Services", 180000)
    row += 1
    _write_row(ws, row, "   40200 Other Income", 12000)
    row += 1
    _write_row(ws, row, "Total Income", 912000, bold_amount=True)
    row += 1

    # --- Cost of Goods Sold ---
    ws.cell(row=row, column=1, value="Cost of Goods Sold").font = BOLD
    row += 1

    _write_row(ws, row, "   50000 Hosting Costs", 86400)
    row += 1
    _write_row(ws, row, "   50100 Support Staff", 120000)
    row += 1
    _write_row(ws, row, "Total Cost of Goods Sold", 206400, bold_amount=True)
    row += 1

    # --- Gross Profit (depth-0 total, orphaned) ---
    _write_row(ws, row, "Gross Profit", 705600, bold_amount=True)
    row += 1

    # --- Expenses ---
    ws.cell(row=row, column=1, value="Expenses").font = BOLD
    row += 1

    _write_row(ws, row, "   60000 Payroll")
    row += 1
    _write_row(ws, row, "      Salaries and Wages", 360000)
    row += 1
    _write_row(ws, row, "      Payroll Taxes", 36000)
    row += 1
    _write_row(ws, row, "      Benefits", 48000)
    row += 1
    _write_row(ws, row, "   Total 60000 Payroll", 444000, bold_amount=True)
    row += 1
    _write_row(ws, row, "   61000 Rent", 72000)
    row += 1
    _write_row(ws, row, "   62000 Software and Subscriptions", 36000)
    row += 1
    _write_row(ws, row, "   63000 Marketing", 60000)
    row += 1
    _write_row(ws, row, "   64000 Legal and Professional", 24000)
    row += 1
    _write_row(ws, row, "   65000 Insurance", 18000)
    row += 1
    _write_row(ws, row, "   66000 Travel", 15000)
    row += 1
    _write_row(ws, row, "   67000 Office Supplies", 6000)
    row += 1
    _write_row(ws, row, "   68000 Meals and Entertainment", 4800)
    row += 1
    _write_row(ws, row, "Total Expenses", 679800, bold_amount=True)
    row += 1

    # --- Net Operating Income (depth-0 summary) ---
    _write_row(ws, row, "Net Operating Income", 25800, bold_amount=True)
    row += 1

    # --- Other Expenses ---
    ws.cell(row=row, column=1, value="Other Expenses").font = BOLD
    row += 1

    _write_row(ws, row, "   70000 Interest Expense", 8400)
    row += 1
    _write_row(ws, row, "   71000 Depreciation", 12000)
    row += 1
    _write_row(ws, row, "Total Other Expenses", 20400, bold_amount=True)
    row += 1

    # --- Net Other Income ---
    _write_row(ws, row, "Net Other Income", -20400, bold_amount=True)
    row += 1

    # --- Net Income ---
    _write_row(ws, row, "Net Income", 5400, bold_amount=True)

    wb.save(str(OUTPUT_PATH))
    print(f"Generated: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    generate()
