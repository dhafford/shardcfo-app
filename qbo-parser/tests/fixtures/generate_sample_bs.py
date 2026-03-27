"""Generate tests/fixtures/sample_bs.xlsx — Balance Sheet for Demo Startup Inc."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

FIXTURES_DIR = Path(__file__).parent
OUTPUT_PATH = FIXTURES_DIR / "sample_bs.xlsx"

BOLD = Font(bold=True)


def _write_header(ws, company: str, title: str, date_range: str) -> None:
    """Write the 3-row QBO header block."""
    ws["A1"] = company
    ws["A1"].font = BOLD
    ws["A2"] = title
    ws["A2"].font = BOLD
    ws["A3"] = date_range
    ws["A3"].font = BOLD
    ws["B5"] = "Total"
    ws["B5"].font = BOLD


def _write_row(ws, row_num: int, label: str, amount=None, bold_amount: bool = False) -> None:
    """Write a single data row.  Label is always bold (QBO convention)."""
    cell_a = ws.cell(row=row_num, column=1, value=label)
    cell_a.font = BOLD
    if amount is not None:
        cell_b = ws.cell(row=row_num, column=2, value=float(amount))
        if bold_amount:
            cell_b.font = BOLD


def generate() -> Path:
    """Create sample_bs.xlsx and return its Path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_header(
        ws,
        company="Demo Startup Inc",
        title="Balance Sheet",
        date_range="As of December 31, 2024",
    )

    row = 6  # first data row

    # --- ASSETS ---
    ws.cell(row=row, column=1, value="ASSETS").font = BOLD
    row += 1

    # Current Assets (depth 1 — section within ASSETS)
    _write_row(ws, row, "   Current Assets")
    row += 1
    _write_row(ws, row, "      10000 Cash and Cash Equivalents", 285000)
    row += 1
    _write_row(ws, row, "      11000 Accounts Receivable", 95000)
    row += 1
    _write_row(ws, row, "      12000 Prepaid Expenses", 18000)
    row += 1
    _write_row(ws, row, "   Total Current Assets", 398000, bold_amount=True)
    row += 1

    # Fixed Assets (depth 1)
    _write_row(ws, row, "   Fixed Assets")
    row += 1
    _write_row(ws, row, "      15000 Equipment", 120000)
    row += 1
    _write_row(ws, row, "      15500 Accumulated Depreciation", -42000)
    row += 1
    _write_row(ws, row, "   Total Fixed Assets", 78000, bold_amount=True)
    row += 1

    _write_row(ws, row, "Total Assets", 476000, bold_amount=True)
    row += 1

    # --- LIABILITIES AND EQUITY ---
    ws.cell(row=row, column=1, value="LIABILITIES AND EQUITY").font = BOLD
    row += 1

    # Liabilities (depth 1)
    _write_row(ws, row, "   Liabilities")
    row += 1

    # Current Liabilities (depth 2)
    _write_row(ws, row, "      Current Liabilities")
    row += 1
    _write_row(ws, row, "         20000 Accounts Payable", 45000)
    row += 1
    _write_row(ws, row, "         21000 Accrued Expenses", 28000)
    row += 1
    _write_row(ws, row, "         22000 Deferred Revenue", 67000)
    row += 1
    _write_row(ws, row, "      Total Current Liabilities", 140000, bold_amount=True)
    row += 1

    # Long-Term Liabilities (depth 2)
    _write_row(ws, row, "      Long-Term Liabilities")
    row += 1
    _write_row(ws, row, "         25000 Notes Payable", 150000)
    row += 1
    _write_row(ws, row, "      Total Long-Term Liabilities", 150000, bold_amount=True)
    row += 1

    _write_row(ws, row, "   Total Liabilities", 290000, bold_amount=True)
    row += 1

    # Equity (depth 1)
    _write_row(ws, row, "   Equity")
    row += 1
    _write_row(ws, row, "      30000 Common Stock", 50000)
    row += 1
    _write_row(ws, row, "      31000 Additional Paid-in Capital", 100000)
    row += 1
    _write_row(ws, row, "      32000 Retained Earnings", 30600)
    row += 1
    _write_row(ws, row, "      Net Income", 5400)
    row += 1
    _write_row(ws, row, "   Total Equity", 186000, bold_amount=True)
    row += 1

    _write_row(ws, row, "Total Liabilities and Equity", 476000, bold_amount=True)

    wb.save(str(OUTPUT_PATH))
    print(f"Generated: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    generate()
