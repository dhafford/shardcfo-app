"""Generate tests/fixtures/sample_pl_monthly.xlsx — Monthly P&L for Demo Startup Inc."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font

FIXTURES_DIR = Path(__file__).parent
OUTPUT_PATH = FIXTURES_DIR / "sample_pl_monthly.xlsx"

BOLD = Font(bold=True)

# 12 monthly column headers as datetime objects (QBO convention)
MONTHS = [datetime(2024, m, 1) for m in range(1, 13)]
# Column labels the parser produces: "Jan 2024", ..., "Dec 2024"
MONTH_LABELS = [dt.strftime("%b %Y") for dt in MONTHS]


def _col_letter(n: int) -> str:
    """1-based column index to letter(s), e.g. 1->'A', 2->'B', 14->'N'."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _saas_monthly() -> List[float]:
    """SaaS Revenue: grows ~5%/month, starting ~45234 so annual sum ≈ 720000.

    All twelve values grow monotonically — January is the lowest, December
    the highest.  The starting value is chosen so the natural 5%-per-month
    compounding produces a sum within a few dollars of 720000.
    """
    values = []
    v = 45234.30  # ~45234 × geometric-sum-factor ≈ 720000
    for _ in range(12):
        values.append(round(v, 2))
        v *= 1.05
    return values


def _professional_services_monthly() -> List[float]:
    """Flat 15000/month = 180000."""
    return [15000.0] * 12


def _hosting_monthly() -> List[float]:
    """Flat 7200/month = 86400."""
    return [7200.0] * 12


def _payroll_monthly() -> List[float]:
    """Flat 37000/month = 444000."""
    return [37000.0] * 12


def _rent_monthly() -> List[float]:
    """Flat 6000/month = 72000."""
    return [6000.0] * 12


def _software_monthly() -> List[float]:
    """Flat 3000/month = 36000."""
    return [3000.0] * 12


def _marketing_monthly() -> List[float]:
    """Growing: start 3500, ends higher, sum = 60000."""
    base = [3500, 4000, 4500, 5000, 5000, 5500, 5500, 5500, 5500, 5500, 5500, 6000]
    s = sum(base[:-1])
    base[-1] = 60000 - s
    return [float(x) for x in base]


def _interest_monthly() -> List[float]:
    """Flat 700/month = 8400."""
    return [700.0] * 12


def generate() -> Path:
    """Create sample_pl_monthly.xlsx and return its Path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Header rows 1-3 ---
    ws["A1"] = "Demo Startup Inc"
    ws["A1"].font = BOLD
    ws["A2"] = "Profit and Loss (Accrual Basis)"
    ws["A2"].font = BOLD
    ws["A3"] = "January - December, 2024"
    ws["A3"].font = BOLD

    # --- Row 5: column headers ---
    # Col A: blank (label column)
    # Cols B..M: datetime objects for Jan-Dec 2024
    # Col N: "Total"
    # Col O: "% of Income"
    header_row = 5
    for i, dt in enumerate(MONTHS):
        col = i + 2  # B=2, C=3, ... M=13
        c = ws.cell(row=header_row, column=col, value=dt)
        c.font = BOLD
    ws.cell(row=header_row, column=14, value="Total").font = BOLD
    ws.cell(row=header_row, column=15, value="% of Income").font = BOLD

    # --- Compute monthly arrays ---
    saas = _saas_monthly()
    prof_svc = _professional_services_monthly()
    hosting = _hosting_monthly()
    payroll = _payroll_monthly()
    rent = _rent_monthly()
    software = _software_monthly()
    marketing = _marketing_monthly()
    interest = _interest_monthly()

    # Derived monthly totals
    total_income_m = [saas[i] + prof_svc[i] for i in range(12)]
    total_cogs_m = [hosting[i] for i in range(12)]
    gross_profit_m = [total_income_m[i] - total_cogs_m[i] for i in range(12)]
    total_expenses_m = [payroll[i] + rent[i] + software[i] + marketing[i] for i in range(12)]
    net_op_income_m = [gross_profit_m[i] - total_expenses_m[i] for i in range(12)]
    total_other_exp_m = [interest[i] for i in range(12)]
    net_income_m = [net_op_income_m[i] - total_other_exp_m[i] for i in range(12)]

    # Annual totals
    total_income_annual = sum(total_income_m)       # 900000
    total_cogs_annual = sum(total_cogs_m)           # 86400
    gross_profit_annual = sum(gross_profit_m)        # 813600
    total_expenses_annual = sum(total_expenses_m)    # 612000
    net_op_income_annual = sum(net_op_income_m)      # 201600
    total_other_exp_annual = sum(total_other_exp_m)  # 8400
    net_income_annual = sum(net_income_m)            # 193200

    def write_data_row(
        row_num: int,
        label: str,
        monthly_vals: List[float],
        annual_total: float,
        pct_str: str = "",
        bold_amounts: bool = False,
    ) -> None:
        """Write label + 12 monthly cols + Total + % of Income."""
        cell_a = ws.cell(row=row_num, column=1, value=label)
        cell_a.font = BOLD
        for i, val in enumerate(monthly_vals):
            c = ws.cell(row=row_num, column=i + 2, value=val)
            if bold_amounts:
                c.font = BOLD
        tc = ws.cell(row=row_num, column=14, value=annual_total)
        if bold_amounts:
            tc.font = BOLD
        if pct_str:
            ws.cell(row=row_num, column=15, value=pct_str)

    def write_header_label(row_num: int, label: str) -> None:
        """Write a section header (no amounts)."""
        ws.cell(row=row_num, column=1, value=label).font = BOLD

    row = 6  # first data row

    # --- Income ---
    write_header_label(row, "Income")
    row += 1

    write_data_row(row, "   40000 SaaS Revenue", saas, sum(saas), "79.00%")
    row += 1
    write_data_row(row, "   40100 Professional Services", prof_svc, sum(prof_svc), "20.00%")
    row += 1
    write_data_row(row, "Total Income", total_income_m, total_income_annual, "100.00%", bold_amounts=True)
    row += 1

    # --- Cost of Goods Sold ---
    write_header_label(row, "Cost of Goods Sold")
    row += 1

    write_data_row(row, "   50000 Hosting Costs", hosting, total_cogs_annual, "9.60%")
    row += 1
    write_data_row(row, "Total Cost of Goods Sold", total_cogs_m, total_cogs_annual, "9.60%", bold_amounts=True)
    row += 1

    # --- Gross Profit ---
    write_data_row(row, "Gross Profit", gross_profit_m, gross_profit_annual, "90.40%", bold_amounts=True)
    row += 1

    # --- Expenses ---
    write_header_label(row, "Expenses")
    row += 1

    write_data_row(row, "   60000 Payroll", payroll, sum(payroll), "49.33%")
    row += 1
    write_data_row(row, "   61000 Rent", rent, sum(rent), "8.00%")
    row += 1
    write_data_row(row, "   62000 Software", software, sum(software), "4.00%")
    row += 1
    write_data_row(row, "   63000 Marketing", marketing, sum(marketing), "6.67%")
    row += 1
    write_data_row(row, "Total Expenses", total_expenses_m, total_expenses_annual, "68.00%", bold_amounts=True)
    row += 1

    # --- Net Operating Income ---
    write_data_row(row, "Net Operating Income", net_op_income_m, net_op_income_annual, "22.40%", bold_amounts=True)
    row += 1

    # --- Other Expenses ---
    write_header_label(row, "Other Expenses")
    row += 1

    write_data_row(row, "   70000 Interest Expense", interest, total_other_exp_annual, "0.93%")
    row += 1
    write_data_row(row, "Total Other Expenses", total_other_exp_m, total_other_exp_annual, "0.93%", bold_amounts=True)
    row += 1

    # --- Net Income ---
    write_data_row(row, "Net Income", net_income_m, net_income_annual, "21.47%", bold_amounts=True)

    wb.save(str(OUTPUT_PATH))
    print(f"Generated: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    generate()
